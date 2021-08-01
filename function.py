import subprocess

from openpyxl import load_workbook, Workbook
from time import strftime
import os

from openpyxl.styles import Alignment
from openpyxl.utils.exceptions import InvalidFileException


# 内网快件查询 代码暂存
# from selenium.webdriver.chrome.options import Options
# from selenium import webdriver
# from selenium.webdriver.common.keys import Keys
# import time
#
# exe_data = r'C:\Users\Hast\AppData\Local\360Chrome\Chrome\Application\360chrome.exe'    #浏览器根目录所在地
# chrome_driver = r"C:\Users\Hast\AppData\Local\360Chrome\Chrome\Application\chromedriver.exe"
#
# chrome_options = Options()
# chrome_options.binary_location =exe_data
# # chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9999")
#
# #
#
# driver = webdriver.Chrome(chrome_driver,chrome_options=chrome_options)
# # driver.get('https://khd.yundasys.com/')
# driver.get('https://idaas.yundasys.com:10443/frontend/login#/login?sp_code=82d6069b1f73569af4d054a4d3d534b6iSMc3qflHGC&application_key=yundaplugin_cas_apereo100&application_disabled=false')
# driver.implicitly_wait(10)
#
# # driver.find_element_by_id("kw").send_keys("xiaobai" + Keys.RETURN)
# # time.sleep(10)
# # driver.quit()
#
# driver.switch_to.frame('customLoginpage')
# driver.find_element_by_xpath('/html/body/div/div[4]/div/div[1]/span[1]').click()
# print("登录")
# driver.find_element_by_xpath('/html/body/div/div[4]/div/div[1]/ul/li[2]').click()
# print("密码登录")
# driver.find_element_by_xpath('/html/body/div/div[4]/div/div[2]/div[2]/form/div[1]/input').send_keys("90086965")
# print("输入账号")
# # driver.find_element_by_xpath('/html/body/div/div[4]/div/div[2]/div[3]/form/div[2]/input').send_keys("Z2stt123456")
# driver.find_element_by_xpath('/html/body/div/div[4]/div/div[2]/div[2]/form/div[2]/input').send_keys("Z2stt123456")
# print("输入密码")
# driver.find_element_by_xpath('/html/body/div/div[4]/div/div[2]/div[2]/form/div[4]/button').click()
# print("登录账号")

def handle_excel(file):
    if file:
        shipid = ""
        scantime = ""
        YT = ""
        try:
            wb = load_workbook(file)  # 打开表格
            ws = wb.active  # 激活这个workbook
            shipid = ws['A']  # 单号
            scantime = ws['B']  # 扫描时间
            YT = ws['F']  # 月台号
            if shipid[0].value != "ship_id":
                return -1
            elif scantime[0].value != "扫描时间":
                return -1
            elif YT[0].value != "月台":
                return -1
        except InvalidFileException:
            return -2
        except Exception as e:
            return e
        TSOD = []  # three super original data
        YTSET = set()  # 月台集合 不重复
        # 单元格类数据转换 对应的int datetime str 三个类型 顺便用这个循环创建了一个月台集合 分割月台的时候要用
        for i in range(len(shipid) - 1):
            TSOD.append([
                shipid[i + 1].value,
                scantime[i + 1].value.strftime("%Y-%m-%d %H:%M:%S"), YT[i + 1].value
            ])
            YTSET.add(YT[i + 1].value)
        # 第三列月台排序    这里排序优化是有必要的 等会要做分割的
        TSOD.sort(key=lambda elem: elem[2])
        # 集合转换列表
        YTSET = list(YTSET)
        # 集合无序 转换列表后优化排序 减少冗余for次数
        YTSET.sort()
        # 确认类型 这句现在无关紧要啦
        print(type(YTSET), YTSET)
        # 列表推导式 用月台集合列表的大小 创建一个空二维列表 方便填充
        YTsingle = [[] for i in range(len(YTSET))]
        # 列表推导式 从原始数据里找到所有的月台 相同的月台号 放到一个列表中 注意YTsingle是一个三维列表
        [
            YTsingle[j].append(TSOD[i]) for j in range(len(YTSET))
            for i in range(len(TSOD)) if TSOD[i][2] == YTSET[j]
        ]
        # 现在对 YTsingle进行每个元素的元素[1]的排序 这里就是排序时间 我先用for解开了一层
        [i.sort(key=lambda elem: elem[1]) for i in YTsingle]
        # final list 最终的列表 把三维列表 合并为二维列表
        FL = [j for i in YTsingle for j in i]
        # 打印测试
        for i in FL:
            print(i)
        return FL
        # 保留了列表推导式之前的写法 以免日后出事 解不开
        # for j in range(len(YTSET)):
        #     for i in range(len(TSOD)):
        #         if TSOD[i][2] == YTSET[j]:
        #             YTsingle[j].append(TSOD[i])


def export_excel(widgetvaluelist, fileaddress):
    # 把列表中的数据做成excel
    print(widgetvaluelist)
    wb = Workbook()
    ws = wb.active
    # 循环列表 坐标
    # for r_index, ODL in enumerate(widgetvaluelist):  # row_index one-dimensional list
    #     for c_index, TDV in enumerate(ODL):  # column_index two-dimensional value
    #         ws.cell(r_index + 1, c_index + 1, str(TDV))
    #         # 这里我没有用单元格格式 内建样式 直接做的字符串转换 举个例子ws['B6'].number_format = 'yyyy-mm-dd'
    #         #     1: '0',       改成没有小数点的0就好了
    #         ws.insert_rows(1)
    #         ws.cell(1,1,["ship_id","扫描时间","月台"])
    ws.append(["ship_id", "扫描时间", "月台"])
    [ws.append(ODL) for ODL in widgetvaluelist]
    for i in ws['A']:
        i.number_format = '0'
        i.alignment = Alignment(horizontal='left', vertical='center')
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    filepath = os.path.join(os.path.expanduser("~"), 'Desktop') + "\\ATO" + fileaddress.split("/")[-1]
    wb.save(filepath)


def ConvertVideosInFoldersToPicture(ffmpegaddress, folderaddress):
    """
    列出FR下的所有文件 检测文件扩展名为dav 开始处理这个文件 -r 1 暂时没有添加-framerate 帧数 这个参数
    :param ffmpegaddress:FFmpeg的文件地址
    :param folderaddress:将要处理的视频所在文件夹
    :return:
    """
    all_file = os.listdir(folderaddress)
    print(ffmpegaddress)
    print(folderaddress)
    for file in all_file:
        shipid = os.path.splitext(file)[0]
        ext = os.path.splitext(file)[1]
        print(file)
        print(shipid)
        print(ext)
        if ext == ".dav":
            parameter = " -i {}\\{} -r 1 {}\\%3d.jpeg".format(videoprocessaddress, file, videoprocessaddress)
            print(parameter)
            cmd = ffmpegaddress + parameter
            print(cmd)
            subprocess.Popen(cmd)

